from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path("test-cases/loan-business-logic-test-cases.xlsx")

headers = [
    "รหัสทดสอบ",
    "หมวด",
    "นศ. เคยกู้",
    "มีหนี้ค้าง",
    "ผู้ดำเนินการ",
    "สถานะก่อนทำ",
    "การกระทำ / ข้อมูลทดสอบ",
    "เงื่อนไขหรือข้อมูลสำคัญ",
    "สถานะที่คาดหวัง",
    "ผู้คืน/ปฏิเสธ",
    "ผลลัพธ์ที่คาดหวังบนเว็บ / API",
    "ความครอบคลุมระบบปัจจุบัน",
    "สรุปเหตุการณ์ทั้งหมด",
    "เสร็จแล้ว",
    "ผลจริง / หมายเหตุ",
]


def case(
    test_id, category, previous, debt, actor, before, action, condition, expected_status,
    decider, expected, coverage, summary,
):
    return [
        test_id, category, previous, debt, actor, before, action, condition, expected_status,
        decider, expected, coverage, summary, "☐", "",
    ]


cases = [
    case("AUTH-001", "สิทธิ์เข้าใช้งาน", "ไม่เคย", "ไม่มี", "ผู้ไม่ล็อกอิน", "-", "ส่งคำขอกู้", "ไม่มี session", "ไม่สร้างคำขอ", "-", "API ตอบ 401 Authentication required", "ทำได้แล้ว", "ผู้ไม่ล็อกอินพยายามยื่นกู้ ระบบไม่อนุญาต",),
    case("AUTH-002", "สิทธิ์เข้าใช้งาน", "ไม่เคย", "ไม่มี", "นักศึกษา", "-", "เปิดคำขอของคนอื่น", "ใช้ loan id ของนักศึกษาคนอื่น", "ไม่แสดงข้อมูล", "-", "API ตอบ 404 และไม่เปิดเผยข้อมูล", "ทำได้แล้ว", "นักศึกษาพยายามเปิดคำขอของผู้อื่น ระบบไม่อนุญาต",),
    case("AUTH-003", "สิทธิ์เข้าใช้งาน", "ไม่เคย", "ไม่มี", "อาจารย์คนอื่น", "pending_advisor", "เปิด/ตัดสินคำขอ", "ไม่ใช่อาจารย์ที่ถูกมอบหมาย", "ไม่เปลี่ยนสถานะ", "-", "API ตอบ 404", "ทำได้แล้ว", "อาจารย์ที่ไม่ได้รับมอบหมายพยายามตัดสินคำขอ ระบบไม่อนุญาต",),
    case("AUTH-004", "สิทธิ์เข้าใช้งาน", "ไม่เคย", "ไม่มี", "บุคลากรทั่วไป", "-", "ดูรายการคำขอ", "ไม่มีบทบาท admin/super_admin/executive/advisor", "ไม่แสดงรายการ", "-", "API ตอบ 403", "ทำได้แล้ว", "ผู้ไม่มีสิทธิ์พยายามดูรายการคำขอ ระบบไม่อนุญาต",),
    case("SUBMIT-001", "ยื่นคำขอ", "ไม่เคย", "ไม่มี", "นักศึกษา", "ไม่มีคำขอเปิด", "ยื่นคำขอถูกต้อง", "จำนวนเงิน > 0, ชั้นปี 1-4, ผ่อน 1-3 งวด, อาจารย์มีเพียง 1 คน", "pending_advisor", "-", "สร้าง LoanRequest, approval advisor attempt 1 และ audit log; หน้าอาจารย์เห็นคำขอ", "ทำได้แล้ว", "นศ.ไม่เคยกู้ ไม่มีหนี้ค้าง ยื่นคำขอถูกต้อง รออาจารย์อนุมัติ",),
    case("SUBMIT-002", "ยื่นคำขอ", "เคย", "ไม่มี", "นักศึกษา", "ไม่มีคำขอเปิด", "ยื่นคำขอถูกต้อง", "มีประวัติคำขอปิดแล้ว (closed)", "pending_advisor", "-", "ยื่นได้ เพราะคำขอ closed ไม่ถือเป็น open loan", "ทำได้แล้ว", "นศ.เคยกู้ ปิดหนี้แล้ว ยื่นคำขอใหม่ รออาจารย์อนุมัติ",),
    case("SUBMIT-003", "ยื่นคำขอ", "ไม่เคย", "มี", "นักศึกษา", "ไม่มีคำขอเปิด", "ยื่นคำขอถูกต้อง", "จำลองว่ามีหนี้ค้างจากแหล่งข้อมูลภายนอก", "pending_advisor", "-", "ระบบปัจจุบันยังไม่มี field/rule ตรวจหนี้ค้าง จึงยังยื่นได้; บันทึกเป็น business-rule gap", "ช่องว่าง: ไม่มี rule หนี้ค้าง", "นศ.ไม่เคยกู้แต่มีหนี้ค้าง ยื่นคำขอ ระบบปัจจุบันยังรออาจารย์อนุมัติ",),
    case("SUBMIT-004", "ยื่นคำขอ", "เคย", "มี", "นักศึกษา", "ไม่มีคำขอเปิด", "ยื่นคำขอถูกต้อง", "เคยกู้และมีหนี้ค้าง", "pending_advisor", "-", "ระบบปัจจุบันยังไม่ตรวจหนี้ค้าง; ทดสอบเพื่อยืนยัน policy ที่ธุรกิจกำหนด", "ช่องว่าง: ไม่มี rule หนี้ค้าง", "นศ.เคยกู้และมีหนี้ค้าง ยื่นคำขอ ระบบปัจจุบันยังรออาจารย์อนุมัติ",),
    case("SUBMIT-005", "ยื่นคำขอ", "เคย", "ไม่มี", "นักศึกษา", "pending_advisor", "ยื่นคำขอใหม่", "มีคำขอ open อยู่", "ไม่สร้างคำขอ", "-", "API ตอบ 409 You already have an open loan request", "ทำได้แล้ว", "นศ.มีคำขอรออาจารย์ ยื่นซ้ำ ระบบป้องกันคำขอเปิดซ้ำ",),
    case("SUBMIT-006", "ยื่นคำขอ", "เคย", "ไม่มี", "นักศึกษา", "pending_admin", "ยื่นคำขอใหม่", "มีคำขอ open อยู่ระหว่างรอเจ้าหน้าที่", "ไม่สร้างคำขอ", "-", "API ตอบ 409", "ทำได้แล้ว", "นศ.ผ่านอาจารย์แล้วแต่ยังรอเจ้าหน้าที่ ยื่นซ้ำ ระบบป้องกันคำขอเปิดซ้ำ",),
    case("SUBMIT-007", "ตรวจสอบข้อมูล", "ไม่เคย", "ไม่มี", "นักศึกษา", "-", "ยื่นจำนวนเงิน 0/ติดลบ/ทศนิยม", "amount ไม่ใช่จำนวนเต็มบวก", "ไม่สร้างคำขอ", "-", "API ตอบ 422 amount is invalid", "ทำได้แล้ว", "นศ.ยื่นจำนวนเงินไม่ถูกต้อง ระบบปฏิเสธข้อมูล",),
    case("SUBMIT-008", "ตรวจสอบข้อมูล", "ไม่เคย", "ไม่มี", "นักศึกษา", "-", "ยื่นชั้นปี 0, 5 หรือทศนิยม", "studentYear ต้องเป็น 1-4", "ไม่สร้างคำขอ", "-", "API ตอบ 422 studentYear is invalid", "ทำได้แล้ว", "นศ.ยื่นชั้นปีไม่ถูกต้อง ระบบปฏิเสธข้อมูล",),
    case("SUBMIT-009", "ตรวจสอบข้อมูล", "ไม่เคย", "ไม่มี", "นักศึกษา", "-", "เลือกผ่อน 0 หรือ 4 งวด", "installmentCount ต้องเป็น 1-3", "ไม่สร้างคำขอ", "-", "API ตอบ 422 installmentCount is invalid", "ทำได้แล้ว", "นศ.เลือกจำนวนงวดไม่ถูกต้อง ระบบปฏิเสธข้อมูล",),
    case("SUBMIT-010", "ตรวจสอบข้อมูล", "ไม่เคย", "ไม่มี", "นักศึกษา", "-", "เว้นข้อมูลบังคับ", "advisorName/purpose/bankName/bankAccountNo/bankAccountName ว่าง", "ไม่สร้างคำขอ", "-", "API ตอบ 422 ระบุ field ที่ invalid", "ทำได้แล้ว", "นศ.กรอกข้อมูลบังคับไม่ครบ ระบบปฏิเสธข้อมูล",),
    case("SUBMIT-011", "ตรวจสอบข้อมูล", "ไม่เคย", "ไม่มี", "นักศึกษา", "-", "เลือกอาจารย์ที่ไม่มี/ชื่อซ้ำ", "พบ advisor 0 คน หรือมากกว่า 1 คน", "ไม่สร้างคำขอ", "-", "API ตอบ 422 advisorName is ambiguous or not found", "ทำได้แล้ว", "นศ.เลือกอาจารย์ไม่ถูกต้องหรือชื่อกำกวม ระบบไม่สร้างคำขอ",),
    case("SUBMIT-012", "ตรวจสอบข้อมูล", "ไม่เคย", "ไม่มี", "นักศึกษา", "-", "ยื่นด้วย CMU identity ไม่ตรง record เดิม", "cmuAccount/email/studentCode ไม่สอดคล้องกัน", "ไม่สร้างคำขอ", "-", "API ตอบ 422 CMU identity does not match the existing student", "ทำได้แล้ว", "นศ.มีข้อมูลระบุตัวตนไม่ตรงกับข้อมูลเดิม ระบบไม่สร้างคำขอ",),
    case("ADVISOR-001", "การพิจารณาอาจารย์", "ไม่เคย", "ไม่มี", "อาจารย์ที่ปรึกษา", "pending_advisor", "อนุมัติ", "approval advisor เป็น pending", "pending_admin", "อาจารย์", "บันทึก approval=approved พร้อมผู้ตัดสิน/เวลา; สร้าง approval admin pending attempt 1", "ทำได้แล้ว", "นศ.ยื่นคำขอ อาจารย์อนุมัติ นศ.ผ่านการอนุมัติโดยอาจารย์และรอเจ้าหน้าที่",),
    case("ADVISOR-002", "การพิจารณาอาจารย์", "ไม่เคย", "ไม่มี", "อาจารย์ที่ปรึกษา", "pending_advisor", "ส่งกลับแก้ไข", "comment ต้องไม่ว่าง", "returned", "อาจารย์", "บันทึก approval=returned พร้อมเหตุผล; นศ.เห็นสถานะ returned", "ทำได้แล้ว", "นศ.ยื่นคำขอ อาจารย์ส่งกลับแก้ไข นศ.ต้องแก้ไขและยื่นใหม่",),
    case("ADVISOR-003", "การพิจารณาอาจารย์", "เคย", "มี", "อาจารย์ที่ปรึกษา", "pending_advisor", "ปฏิเสธ", "comment ต้องไม่ว่าง", "rejected", "อาจารย์", "บันทึก approval=rejected พร้อมเหตุผล; คำขอไม่ใช่ open loan อีกต่อไป", "ทำได้แล้ว", "นศ.เคยกู้และมีหนี้ค้าง อาจารย์ปฏิเสธคำขอพร้อมเหตุผล",),
    case("ADVISOR-004", "การพิจารณาอาจารย์", "ไม่เคย", "ไม่มี", "อาจารย์ที่ปรึกษา", "pending_advisor", "ส่งกลับ/ปฏิเสธโดยไม่ใส่เหตุผล", "comment ว่าง", "ไม่เปลี่ยนสถานะ", "-", "API ตอบ 422 A comment is required for this decision", "ทำได้แล้ว", "อาจารย์พยายามส่งกลับหรือปฏิเสธโดยไม่ให้เหตุผล ระบบไม่อนุญาต",),
    case("ADVISOR-005", "การพิจารณาอาจารย์", "ไม่เคย", "ไม่มี", "อาจารย์ที่ปรึกษา", "pending_admin", "ตัดสินซ้ำ", "คำขอไม่อยู่ pending_advisor แล้ว", "ไม่เปลี่ยนสถานะ", "-", "API ตอบ 409 Loan request has already been decided", "ทำได้แล้ว", "นศ.ผ่านอาจารย์แล้ว อาจารย์พยายามตัดสินซ้ำ ระบบป้องกันข้อมูลซ้ำ",),
    case("RESUBMIT-001", "แก้ไขและยื่นใหม่", "ไม่เคย", "ไม่มี", "นักศึกษา", "returned", "แก้ไขและยื่นใหม่", "คำขอถูกส่งกลับจากอาจารย์ และเลือกอาจารย์เดิม", "pending_advisor", "-", "อัปเดตข้อมูลคำขอ, reset approvedAmount, สร้าง approval advisor attempt ใหม่", "ทำได้แล้ว", "นศ.ถูกอาจารย์ส่งกลับ แก้ไขและยื่นให้อาจารย์เดิมพิจารณาใหม่",),
    case("RESUBMIT-002", "แก้ไขและยื่นใหม่", "ไม่เคย", "ไม่มี", "นักศึกษา", "returned", "แก้ไขและยื่นใหม่", "คำขอถูกส่งกลับจากอาจารย์ และเปลี่ยนอาจารย์", "pending_advisor", "-", "เริ่ม flow ที่อาจารย์ใหม่ โดยสร้าง approval advisor attempt ใหม่", "ทำได้แล้ว", "นศ.ถูกอาจารย์ส่งกลับ เปลี่ยนอาจารย์และยื่นให้อาจารย์ใหม่พิจารณา",),
    case("RESUBMIT-003", "แก้ไขและยื่นใหม่", "ไม่เคย", "ไม่มี", "นักศึกษา", "returned", "แก้ไขและยื่นใหม่", "คำขอถูกส่งกลับจากเจ้าหน้าที่ และเลือกอาจารย์เดิม", "pending_admin", "เจ้าหน้าที่", "ข้ามอาจารย์และสร้าง approval admin attempt ใหม่", "ทำได้แล้วตาม route", "นศ.ผ่านอาจารย์ ถูกเจ้าหน้าที่ส่งกลับ แก้ไขแล้วรอเจ้าหน้าที่ตรวจซ้ำ",),
    case("RESUBMIT-004", "แก้ไขและยื่นใหม่", "ไม่เคย", "ไม่มี", "นักศึกษา", "returned", "แก้ไขและยื่นใหม่", "คำขอถูกส่งกลับจากเจ้าหน้าที่ แต่เปลี่ยนอาจารย์", "pending_advisor", "เจ้าหน้าที่", "กลับไปเริ่มที่อาจารย์ใหม่", "ทำได้แล้วตาม route", "นศ.ถูกเจ้าหน้าที่ส่งกลับแต่เปลี่ยนอาจารย์ จึงกลับไปรออาจารย์อนุมัติ",),
    case("RESUBMIT-005", "แก้ไขและยื่นใหม่", "ไม่เคย", "ไม่มี", "นักศึกษา", "pending_advisor", "เรียกยื่นใหม่", "สถานะไม่ใช่ returned", "ไม่เปลี่ยนสถานะ", "-", "API ตอบ 404 Loan request not found", "ทำได้แล้ว", "นศ.พยายามยื่นใหม่ทั้งที่คำขอยังไม่ถูกส่งกลับ ระบบไม่อนุญาต",),
    case("ADMIN-001", "การตรวจสอบเจ้าหน้าที่", "ไม่เคย", "ไม่มี", "เจ้าหน้าที่", "pending_admin", "อนุมัติ", "คำขอผ่านอาจารย์", "pending_executive", "เจ้าหน้าที่", "ควรบันทึก approval admin=approved และส่งต่อผู้บริหาร", "ช่องว่าง: ยังไม่มี API/หน้าจอ admin decision", "นศ.ผ่านการอนุมัติโดยอาจารย์ ผ่านการตรวจสอบโดยเจ้าหน้าที่ และรอผู้บริหาร",),
    case("ADMIN-002", "การตรวจสอบเจ้าหน้าที่", "ไม่เคย", "ไม่มี", "เจ้าหน้าที่", "pending_admin", "ส่งกลับแก้ไข", "ต้องใส่เหตุผล", "returned", "เจ้าหน้าที่", "ควรบันทึกเหตุผลและเปิดให้นศ.แก้ไข/ยื่นใหม่", "ช่องว่าง: ยังไม่มี API/หน้าจอ admin decision", "นศ.ผ่านอาจารย์ แต่ถูกเจ้าหน้าที่ส่งกลับแก้ไข",),
    case("ADMIN-003", "การตรวจสอบเจ้าหน้าที่", "เคย", "มี", "เจ้าหน้าที่", "pending_admin", "ปฏิเสธ", "พบหนี้ค้างตาม policy", "rejected", "เจ้าหน้าที่", "ควรบันทึกผู้ปฏิเสธและเหตุผล; นศ.เห็นผลชัดเจน", "ช่องว่าง: ยังไม่มี API/หน้าจอ admin decision และ rule หนี้ค้าง", "นศ.ผ่านอาจารย์ แต่ไม่ผ่านการตรวจสอบโดยเจ้าหน้าที่เพราะมีหนี้ค้าง",),
    case("EXEC-001", "การอนุมัติผู้บริหาร", "ไม่เคย", "ไม่มี", "ผู้บริหาร", "pending_executive", "อนุมัติ", "คำขอผ่านอาจารย์และเจ้าหน้าที่", "pending_disbursement", "ผู้บริหาร", "ควรบันทึก approval executive=approved และส่งต่อเพื่อโอนเงิน", "ช่องว่าง: ยังไม่มี API/หน้าจอ executive decision", "นศ.ผ่านอาจารย์และเจ้าหน้าที่ ผ่านการอนุมัติโดยผู้บริหาร และรอโอนเงิน",),
    case("EXEC-002", "การอนุมัติผู้บริหาร", "ไม่เคย", "ไม่มี", "ผู้บริหาร", "pending_executive", "ส่งกลับแก้ไข", "ต้องใส่เหตุผล", "returned", "ผู้บริหาร", "ควรบันทึกเหตุผลและกำหนดเส้นทางการส่งกลับให้ชัดเจน", "ช่องว่าง: ยังไม่มี API/หน้าจอ executive decision", "นศ.ผ่านอาจารย์และเจ้าหน้าที่ แต่ถูกผู้บริหารส่งกลับแก้ไข",),
    case("EXEC-003", "การอนุมัติผู้บริหาร", "เคย", "มี", "ผู้บริหาร", "pending_executive", "ปฏิเสธ", "พบเหตุผลตาม policy", "rejected", "ผู้บริหาร", "ควรบันทึกผู้ปฏิเสธและเหตุผล; นศ.เห็นผลชัดเจน", "ช่องว่าง: ยังไม่มี API/หน้าจอ executive decision", "นศ.ผ่านอาจารย์และเจ้าหน้าที่ แต่ไม่ผ่านการอนุมัติโดยผู้บริหาร",),
    case("DISBURSE-001", "การโอนเงิน", "ไม่เคย", "ไม่มี", "เจ้าหน้าที่การเงิน", "pending_disbursement", "บันทึกการโอนสำเร็จ", "ยอด/บัญชีผู้รับถูกต้อง", "disbursed", "-", "ควรบันทึก disbursedAt, FundTransaction และแสดงกำหนดชำระ", "ช่องว่าง: ยังไม่มี API/หน้าจอโอนเงิน", "นศ.ผ่านทุกขั้นตอน เจ้าหน้าที่โอนเงินสำเร็จ นศ.ได้รับเงินกู้",),
    case("DISBURSE-002", "การโอนเงิน", "ไม่เคย", "ไม่มี", "เจ้าหน้าที่การเงิน", "pending_disbursement", "โอนไม่สำเร็จ", "บัญชีไม่ถูกต้อง/ระบบธนาคารขัดข้อง", "pending_disbursement", "-", "ควรไม่เปลี่ยนเป็น disbursed และเก็บเหตุผล/หลักฐานการลองโอน", "ช่องว่าง: ยังไม่มี API/หน้าจอโอนเงิน", "นศ.ผ่านทุกขั้นตอน แต่การโอนเงินไม่สำเร็จ คำขอยังคงรอโอน",),
    case("REPAY-001", "การชำระหนี้", "เคย", "มี", "นักศึกษา", "disbursed", "ชำระบางส่วน", "ยอดชำระน้อยกว่ายอดครบทุกงวด", "disbursed", "-", "ควรเพิ่ม Payment และอัปเดต amountPaid โดยยังมีหนี้ค้าง", "ช่องว่าง: schema รองรับ แต่ยังไม่มี API ยืนยันชำระ", "นศ.ได้รับเงินกู้ ชำระบางส่วน ยังคงมีหนี้ค้าง",),
    case("REPAY-002", "การชำระหนี้", "เคย", "มี", "เจ้าหน้าที่", "disbursed", "ยืนยันชำระครบ", "ทุก installment ชำระครบ", "closed", "-", "ควรระบุ settledAt ของงวดและ closedAt ของคำขอ", "ช่องว่าง: ยังไม่มี API/logic ปิดหนี้", "นศ.ชำระครบทุกงวด เจ้าหน้าที่ตรวจสอบแล้ว ปิดหนี้เรียบร้อย",),
    case("CANCEL-001", "ยกเลิก", "ไม่เคย", "ไม่มี", "นักศึกษา/เจ้าหน้าที่", "draft หรือ pending_*", "ยกเลิกคำขอ", "ต้องบันทึกผู้ยกเลิกและเวลา", "cancelled", "ผู้ยกเลิก", "ควรบันทึก cancelledAt/cancelledBy และอนุญาตให้ยื่นใหม่ได้", "ช่องว่าง: schema รองรับ แต่ยังไม่มี API ยกเลิก", "นศ.ยกเลิกคำขอก่อนรับเงิน คำขอถูกยกเลิกและยื่นใหม่ได้",),
    case("DISPLAY-001", "การแสดงผล", "ไม่เคย", "ไม่มี", "นักศึกษา", "pending_advisor", "ดู dashboard/รายละเอียดคำขอ", "มี approval advisor pending", "pending_advisor", "-", "หน้าเว็บต้องแสดงว่ารออาจารย์ ไม่แสดงว่าอนุมัติแล้ว", "ทดสอบ UI เทียบ API", "นศ.ยื่นคำขอแล้ว หน้าเว็บแสดงสถานะรออาจารย์อนุมัติถูกต้อง",),
    case("DISPLAY-002", "การแสดงผล", "ไม่เคย", "ไม่มี", "นักศึกษา", "pending_admin", "ดู dashboard/รายละเอียดคำขอ", "advisor approved; admin pending", "pending_admin", "-", "หน้าเว็บต้องสรุปว่า “ผ่านอาจารย์, รอเจ้าหน้าที่” และชื่อผู้อนุมัติถูกต้อง", "ทดสอบ UI เทียบ API", "นศ.ผ่านการอนุมัติโดยอาจารย์ และรอการตรวจสอบโดยเจ้าหน้าที่",),
    case("DISPLAY-003", "การแสดงผล", "เคย", "มี", "นักศึกษา", "rejected", "ดูประวัติคำขอ", "rejected โดยอาจารย์/เจ้าหน้าที่/ผู้บริหาร", "rejected", "ผู้ตัดสินจริง", "หน้าเว็บต้องแสดงขั้นตอนและผู้ปฏิเสธจริง พร้อมเหตุผล", "ทดสอบ UI เทียบ API", "นศ.ไม่ผ่านการอนุมัติ หน้าเว็บแสดงว่าโดนปฏิเสทโดยใครและเหตุผลถูกต้อง",),
    case("DISPLAY-004", "การแสดงผล", "เคย", "มี", "นักศึกษา", "returned", "ดูคำขอที่ถูกส่งกลับ", "returned พร้อม comment", "returned", "ผู้ส่งกลับจริง", "หน้าเว็บต้องแสดงเหตุผลและเปิดทางแก้ไข/ยื่นใหม่เฉพาะเจ้าของคำขอ", "ทดสอบ UI เทียบ API", "นศ.ถูกส่งกลับแก้ไข หน้าเว็บแสดงเหตุผลและให้ยื่นใหม่ได้",),
]


def style_sheet(ws, freeze="A2"):
    navy = "1F4E78"
    blue = "D9EAF7"
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for row_num in range(2, ws.max_row + 1):
        if row_num % 2 == 0:
            for cell in ws[row_num]:
                cell.fill = PatternFill("solid", fgColor="F7FBFF")
    return blue


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "สรุป"
    tests = wb.create_sheet("รายการทดสอบ")
    statuses = wb.create_sheet("สถานะและกติกา")

    # Test-case sheet
    tests.append(headers)
    for row in cases:
        tests.append(row)
    style_sheet(tests)
    widths = [14, 20, 12, 12, 18, 19, 31, 38, 22, 18, 46, 31, 54, 13, 30]
    for index, width in enumerate(widths, start=1):
        tests.column_dimensions[chr(64 + index)].width = width
    tests.row_dimensions[1].height = 42
    for row_num in range(2, tests.max_row + 1):
        tests.row_dimensions[row_num].height = 62
    done_validation = DataValidation(type="list", formula1='"☐,☑"', allow_blank=False)
    done_validation.promptTitle = "สถานะการทดสอบ"
    done_validation.prompt = "เลือก ☑ เมื่อทดสอบเสร็จแล้ว"
    tests.add_data_validation(done_validation)
    done_validation.add(f"N2:N{tests.max_row}")
    tests.conditional_formatting.add(
        f"N2:N{tests.max_row}",
        CellIsRule(operator="equal", formula=['"☑"'], fill=PatternFill("solid", fgColor="C6E0B4")),
    )

    # Summary sheet
    summary_rows = [
        ["เอกสารทดสอบ Business Logic ระบบกู้ยืมฉุกเฉิน"],
        ["วัตถุประสงค์", "ใช้ตรวจความสอดคล้องระหว่างหน้าเว็บ/API/ฐานข้อมูลกับกติกาธุรกิจ"],
        ["วิธีใช้", "กรองรายการตามหมวดหรือสถานะ แล้วเลือก ☑ ในคอลัมน์ ‘เสร็จแล้ว’ หลังทดสอบจริง"],
        ["หมายเหตุสำคัญ", "‘เคยกู้’ อ้างอิงประวัติคำขอ และ ‘มีหนี้ค้าง’ ยังไม่มี field/rule บังคับในโค้ดปัจจุบัน"],
        [],
        ["ตัวชี้วัด", "จำนวน"],
        ["จำนวน test cases ทั้งหมด", "=COUNTA('รายการทดสอบ'!A2:A1000)"],
        ["ทดสอบเสร็จแล้ว", "=COUNTIF('รายการทดสอบ'!N2:N1000,\"☑\")"],
        ["ยังไม่ทดสอบ", "=COUNTIF('รายการทดสอบ'!N2:N1000,\"☐\")"],
        ["กรณีที่เป็นช่องว่างของระบบ", "=COUNTIF('รายการทดสอบ'!L2:L1000,\"ช่องว่าง*\")"],
        [],
        ["คำอธิบายสถานะ", "draft: ร่าง | returned: ส่งกลับแก้ไข | pending_advisor: รออาจารย์ | pending_admin: รอเจ้าหน้าที่ | pending_executive: รอผู้บริหาร | pending_disbursement: รอโอน | disbursed: โอนแล้ว | closed: ปิดหนี้ | rejected: ปฏิเสธ | cancelled: ยกเลิก"],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.merge_cells("A1:B1")
    summary["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.row_dimensions[1].height = 30
    for row in range(2, summary.max_row + 1):
        summary[f"A{row}"].font = Font(name="Arial", bold=True, size=10)
        summary[f"A{row}"].fill = PatternFill("solid", fgColor="D9EAF7")
        for col in ("A", "B"):
            summary[f"{col}{row}"].font = Font(name="Arial", size=10, bold=(row >= 6 and row <= 10))
            summary[f"{col}{row}"].alignment = Alignment(vertical="top", wrap_text=True)
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 120
    for row in range(2, summary.max_row + 1):
        summary.row_dimensions[row].height = 34
    summary.freeze_panes = "A2"

    # Rules sheet
    rule_rows = [
        ["สถานะ", "ความหมาย", "ผู้ดำเนินการที่คาดหวัง", "หลักฐานในระบบปัจจุบัน"],
        ["pending_advisor", "รออาจารย์ที่ปรึกษาตัดสิน", "อาจารย์ที่ถูกกำหนดใน advisorId", "มี API และ query ตัดสิน"],
        ["pending_admin", "ผ่านอาจารย์และรอเจ้าหน้าที่", "เจ้าหน้าที่/admin", "สร้าง approval admin ได้ แต่ยังไม่มี API ตัดสิน"],
        ["pending_executive", "ผ่านเจ้าหน้าที่และรอผู้บริหาร", "ผู้บริหาร", "มี enum/schema แต่ยังไม่มี flow"],
        ["pending_disbursement", "ผ่านผู้บริหารและรอโอนเงิน", "เจ้าหน้าที่การเงิน", "มี enum/schema แต่ยังไม่มี flow"],
        ["returned", "ส่งกลับให้นักศึกษาแก้ไขและยื่นใหม่", "ผู้ตัดสินของ step ที่ส่งกลับ", "route resubmit รองรับการเลือกเส้นทาง advisor/admin"],
        ["rejected", "คำขอถูกปฏิเสธและไม่ถือเป็นคำขอเปิด", "อาจารย์/เจ้าหน้าที่/ผู้บริหาร", "อาจารย์ทำได้แล้ว; ขั้นอื่นยังไม่มี API"],
        ["disbursed", "โอนเงินกู้แล้ว", "เจ้าหน้าที่การเงิน", "มี enum/schema แต่ยังไม่มี API"],
        ["closed", "ชำระครบและปิดหนี้", "ระบบ/เจ้าหน้าที่", "มี enum/schema แต่ยังไม่มี logic ปิดหนี้"],
        ["cancelled", "ยกเลิกคำขอ", "นักศึกษา/เจ้าหน้าที่", "มี enum/schema แต่ยังไม่มี API"],
        [],
        ["กติกาปัจจุบัน", "รายละเอียด", "", ""],
        ["คำขอเปิดได้เพียง 1 รายการ", "ห้ามมี loan status ที่ไม่ใช่ closed/rejected/cancelled มากกว่า 1 รายการต่อคน", "", ""],
        ["การตัดสินอาจารย์", "approved → pending_admin; returned/rejected ต้องมี comment", "", ""],
        ["การยื่นใหม่", "ทำได้เฉพาะ status=returned; หากถูก admin ส่งกลับและไม่เปลี่ยนอาจารย์ จะกลับ pending_admin", "", ""],
        ["หนี้ค้าง", "โครงสร้างปัจจุบันไม่มีตัวบ่งชี้/กติกาบล็อกหนี้ค้าง จึงต้องยืนยัน policy เพิ่ม", "", ""],
    ]
    for row in rule_rows:
        statuses.append(row)
    style_sheet(statuses)
    for col, width in {"A": 28, "B": 70, "C": 34, "D": 52}.items():
        statuses.column_dimensions[col].width = width
    for row in range(2, statuses.max_row + 1):
        statuses.row_dimensions[row].height = 40

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
